import json
from fastapi import FastAPI, Query, Request, Form, HTTPException, UploadFile, Depends
from fastapi.params import File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from app.db.mongodb import add_lecture, get_user_by_login, get_user_progress, generate_block_id, generate_task_id,generate_lecture_id, get_lecture_by_id, get_all_lectures
from pymongo import MongoClient
from code_test.code_tester import CodeTesterManager
import logging
from datetime import datetime
from app.db.mongodb import update_user_progress
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import load_workbook
import io
from user_generator.credentials_generator import generate_credentials
from typing import List
import signal
import asyncio

# Константы для ограничений
MIN_LOGIN_LENGTH = 4
MAX_LOGIN_LENGTH = 20
MIN_PASSWORD_LENGTH = 6
MAX_PASSWORD_LENGTH = 30

app = FastAPI()

# Настройка Jinja2 
templates = Jinja2Templates(directory="templates")

# Подключение к MongoDB
client = MongoClient("mongodb://localhost:27017/")
db = client["kvant_users"]

def get_db():
    """Функция для внедрения зависимости базы данных"""
    try:
        yield db
    finally:
        pass  # Закрытие клиента будет происходить при завершении приложения

# Создаем коллекции с использованием зависимости
def get_collections(db=Depends(get_db)):
    return {
        "blocks": db["blocks"],
        "users": db["users"],
        "user_progress": db["user_progress"],
        "lectures": db["lectures"]
    }

# Глобальные переменные для обратной совместимости
blocks_collection = db["blocks"]
users_collection = db["users"]
user_progress_collection = db["user_progress"]
lectures_collection = db["lectures"]

# Создаем менеджер для CodeTester
code_tester_manager = CodeTesterManager()

# Логирование
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Добавляем CORS middleware для обработки параллельных запросов
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.add_middleware(
    SessionMiddleware,
    secret_key="kvant_top",
    session_cookie="session_cookie"
)

@app.on_event("startup")
async def startup_event():
    """Инициализация при запуске приложения"""
    logger.info("Запуск приложения")
    # Здесь можно добавить дополнительную инициализацию

@app.on_event("shutdown")
async def shutdown_event():
    """Очистка ресурсов при завершении работы приложения"""
    try:
        logger.info("Начало завершения работы приложения")
        
        # Закрываем соединение с MongoDB
        client.close()
        logger.info("MongoDB соединение закрыто")
        
        # Завершаем работу CodeTesterManager
        if code_tester_manager is not None:
            try:
                await code_tester_manager.shutdown()
                logger.info("CodeTesterManager успешно завершен")
            except Exception as e:
                logger.error(f"Ошибка при завершении CodeTesterManager: {str(e)}")
        
        logger.info("Завершение работы приложения выполнено успешно")
    except Exception as e:
        logger.error(f"Ошибка при завершении работы приложения: {str(e)}", exc_info=True)

def handle_exit(signum, frame):
    """Обработчик сигналов завершения"""
    logger.info(f"Получен сигнал завершения: {signum}")
    raise KeyboardInterrupt()

# Маршрут для отображения страницы авторизации
@app.get("/", response_class=HTMLResponse)
async def read_login_page(request: Request):
    logger.info("Отображение страницы авторизации")
    return templates.TemplateResponse("login.html", {"request": request})

# Маршрут для обработки данных из формы
@app.post("/login", response_class=HTMLResponse)
async def login(request: Request, login: str = Form(...), password: str = Form(...)):
    user = users_collection.find_one({"login": login, "password": password})
    
    if user and user['password'] == password:
        # Сохраняем login в сессии
        request.session["login"] = login
        logger.info(f"Пользователь {login} успешно авторизован")
        return RedirectResponse(url="/dashboard", status_code=303)
    else:
        logger.warning(f"Неверный логин или пароль для пользователя {login}")
        return templates.TemplateResponse("login.html", {"request": request, "error": "Неверный логин или пароль"})

# Роут для dashboard
@app.get("/dashboard", response_class=HTMLResponse)
async def read_dashboard_page(request: Request):
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
   
    user = users_collection.find_one({"login": login})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    logger.info(f"Пользователь {login} открыл dashboard")
    
    blocks = list(blocks_collection.find({}))
    
    # Сортируем блоки по block_id
    blocks.sort(key=lambda x: x["block_id"])
    
    # Получаем прогресс пользователя (асинхронно)
    user_progress = get_user_progress(login)
    logger.info("User Progress: %s", user_progress)
    
    # Создаем словарь для хранения статусов заданий
    task_status = {}
    for progress in user_progress:
        key = (progress["block_id"], progress["task_id"])
        task_status[key] = progress["status"]
    
    # Разделяем задания на три группы
    not_started = []
    in_progress = []
    completed = []
    
    for block in blocks:
        if "tasks" in block:
            # Сортируем задания по task_id
            block["tasks"].sort(key=lambda x: x["task_id"])
             
            for task in block["tasks"]:
                if "task_id" in task:
                    # Формируем ключ для поиска статуса
                    key = (block["block_id"], task["task_id"])
                    status = task_status.get(key, "not_started")  # По умолчанию "not_started"
                    
                    task_info = {
                        "block_id": block["block_id"],
                        "block_name": block["block_name"],
                        "task_id": task["task_id"],
                        "task_name": task["task_name"],
                        "task_description": task["task_description"],
                        "status": status
                    }
                    
                    if status == "not_started":
                        not_started.append(task_info)
                    elif status == "in_progress":
                        in_progress.append(task_info)
                    elif status == "completed":
                        completed.append(task_info)
                else:
                    logger.warning(f"Задание в блоке {block['block_id']} не содержит ключ 'task_id'")
        else:
            logger.warning(f"Блок {block['block_id']} не содержит ключ 'tasks'")
    
    # Получаем все лекции из MongoDB
    lectures = list(lectures_collection.find({}))
    logger.info("Not Started Tasks: %s", not_started)
    logger.info("In Progress Tasks: %s", in_progress)
    logger.info("Completed Tasks: %s", completed)
    
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "user": user,
            "lectures": lectures,  # Передаем список лекций
            "not_started": not_started,  # Не начатые задания
            "in_progress": in_progress,  # Задания в процессе
            "completed": completed,  # Завершенные задания
        }
    )
    
# Маршрут для отображения страницы задания
# Роут для отображения страницы задания (GET)
@app.get("/task/{block_id}/{task_id}", response_class=HTMLResponse)
async def show_task(
    request: Request,
    block_id: int,
    task_id: int,
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Получаем задание из MongoDB
    task = blocks_collection.find_one(
        {"block_id": block_id, "tasks.task_id": task_id},
        {"tasks.$": 1, "block_name": 1}
    )
    if not task:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    
    current_task = task["tasks"][0]
    
    # Получаем прогресс пользователя для этого задания
    user_progress = user_progress_collection.find_one(
        {"login": login, "block_id": block_id, "task_id": task_id}
    )
    status = user_progress["status"] if user_progress else "not_started"
    
    return templates.TemplateResponse(
        "task.html",
        {
            "request": request,
            "block_id": block_id,  # Добавляем block_id в контекст
            "block_name": task["block_name"],
            "task_name": current_task["task_name"],
            "task_description": current_task["task_description"],
            "task_id": current_task["task_id"],
            "test_cases": current_task["test_cases"],
            "is_superuser": user.get("group") == "superuser",
            "status": status,
            "login": login,  # Добавляем login в контекст
        }
    )

# Роут для обработки отправки задания (POST)
@app.post("/task/{block_id}/{task_id}", response_class=HTMLResponse)
async def submit_task(
    request: Request,
    block_id: int,
    task_id: int,
    code: str = Form(...),
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Получаем задание из MongoDB
    task = blocks_collection.find_one(
        {"block_id": block_id, "tasks.task_id": task_id},
        {"tasks.$": 1, "block_name": 1}
    )
    if not task:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    
    current_task = task["tasks"][0]
    
    try:
        # Генерируем уникальный идентификатор для запроса
        request_id = f"{login}_{block_id}_{task_id}_{datetime.now().timestamp()}"
        
        try:
            # Получаем экземпляр CodeTester для текущего запроса
            tester = await code_tester_manager.get_tester(request_id)
            
            # Тестируем код
            results = await tester.test_code(code, current_task["test_cases"])
            
            # Определяем статус задания
            all_passed = all(result["passed"] for result in results)
            status = "completed" if all_passed else "in_progress"
            
            # Форматируем результаты тестов для лучшей читаемости
            formatted_results = []
            for i, result in enumerate(results, 1):
                test_case = current_task["test_cases"][i-1]
                formatted_result = {
                    "passed": result["passed"],
                    "test_number": i,
                    "input": test_case["input"],
                    "expected": test_case["expected_output"],
                    "actual": result.get("actual_output", ""),
                    "error": result.get("error", ""),
                    "message": result.get("message", "")
                }
                formatted_results.append(formatted_result)
            
            # Обновляем прогресс пользователя
            user_progress_collection.update_one(
                {"login": login, "block_id": block_id, "task_id": task_id},
                {"$set": {
                    "status": status,
                    "last_attempt": datetime.now(),
                    "last_code": code
                }},
                upsert=True
            )
            
            return templates.TemplateResponse(
                "task.html",
                {
                    "request": request,
                    "block_id": block_id,
                    "block_name": task["block_name"],
                    "task_name": current_task["task_name"],
                    "task_description": current_task["task_description"],
                    "task_id": current_task["task_id"],
                    "test_cases": current_task["test_cases"],
                    "is_superuser": user.get("group") == "superuser",
                    "results": formatted_results,
                    "code": code,
                    "status": status,
                    "all_passed": all_passed
                }
            )
        finally:
            # Очищаем ресурсы после выполнения
            await code_tester_manager.cleanup_user(request_id)
            
    except Exception as e:
        logger.error(f"Ошибка при тестировании кода: {str(e)}", exc_info=True)
        return templates.TemplateResponse(
            "task.html",
            {
                "request": request,
                "block_id": block_id,
                "block_name": task["block_name"],
                "task_name": current_task["task_name"],
                "task_description": current_task["task_description"],
                "task_id": current_task["task_id"],
                "test_cases": current_task["test_cases"],
                "is_superuser": user.get("group") == "superuser",
                "error": f"Ошибка при выполнении кода: {str(e)}",
                "code": code,
                "status": "in_progress"
            }
        )

@app.get("/task/{block_id}/{task_id}/edit", response_class=HTMLResponse)
async def edit_task_page(
    request: Request,
    block_id: int,
    task_id: int,
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user or user.get("group") != "superuser":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    
    # Получаем задание из MongoDB
    task = blocks_collection.find_one(
        {"block_id": block_id, "tasks.task_id": task_id},
        {"tasks.$": 1, "block_name": 1}
    )
    if not task:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    
    current_task = task["tasks"][0]
    
    return templates.TemplateResponse(
        "edit_task.html",
        {
            "request": request,
            "block_id": block_id,
            "task_id": task_id,
            "block_name": task["block_name"],
            "task_name": current_task["task_name"],
            "task_description": current_task["task_description"],
            "test_cases": current_task["test_cases"],
        }
    )

# Роут для обработки данных формы редактирования задания
@app.post("/task/{block_id}/{task_id}/edit", response_class=RedirectResponse)
async def edit_task(
    request: Request,
    block_id: int,
    task_id: int,
    task_name: str = Form(...),
    task_description: str = Form(...),
    input: list[str] = Form(...),  # Входные данные
    output: list[str] = Form(...),  # Ожидаемые выходные данные
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user or user.get("group") != "superuser":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    
    # Преобразуем входные и выходные данные в тест-кейсы
    test_cases = []
    for i in range(len(input)):
        test_cases.append({
            "input": input[i].strip(),  # Убираем лишние пробелы
            "expected_output": output[i].strip(),  # Убираем лишние пробелы
        })

    # Обновляем задание
    blocks_collection.update_one(
        {"block_id": block_id, "tasks.task_id": task_id},
        {"$set": {
            "tasks.$.task_name": task_name,
            "tasks.$.task_description": task_description,
            "tasks.$.test_cases": test_cases,
        }}
    )
    
    # Перенаправляем пользователя на страницу задания
    return RedirectResponse(url=f"/task/{block_id}/{task_id}", status_code=303)

# Роут для удаления задания
@app.post("/task/{block_id}/{task_id}/delete", response_class=RedirectResponse)
async def delete_task(
    request: Request,
    block_id: int,
    task_id: int,
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user or user.get("group") != "superuser":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    
    # Удаляем задание из блока
    blocks_collection.update_one(
        {"block_id": block_id},
        {"$pull": {"tasks": {"task_id": task_id}}}
    )
    
    # Перенаправляем пользователя на dashboard
    return RedirectResponse(url="/dashboard", status_code=303)


@app.get("/students", response_class=HTMLResponse)
async def view_students(request: Request):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user or user.get("group") != "superuser":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    
    # Получаем всех учеников (исключая суперпользователя)
    students = list(users_collection.find({"group": {"$ne": "superuser"}}))
    
    # Получаем прогресс каждого ученика
    students_with_progress = []
    for student in students:
        progress = list(user_progress_collection.find({"login": student["login"]}))
        students_with_progress.append({
            "last_name": student["last_name"],
            "first_name": student["first_name"],
            "group": student["group"],
            "login": student["login"],
            "password": student["password"],
            "progress": progress
        })
    
    # Получаем список всех уникальных групп
    groups = users_collection.distinct("group", {"group": {"$ne": "superuser"}})
    
    return templates.TemplateResponse(
        "students.html",
        {
            "request": request,
            "students": students_with_progress,
            "groups": groups  # Передаем список групп в шаблон
        }
    )
    
# Маршрут для отображения страницы добавления задания
@app.get("/add_task_page", response_class=HTMLResponse)
async def add_task_page(request: Request):
    return templates.TemplateResponse("add_task.html", {"request": request})

# Маршрут для обработки данных формы и добавления задания
@app.post("/add_task", response_class=RedirectResponse)
async def add_task(
    request: Request,
    task_name: str = Form(...),
    task_description: str = Form(...),
    input: list[str] = Form(...),  # Входные данные
    output: list[str] = Form(...),  # Ожидаемые выходные данные
):
    # Преобразуем входные и выходные данные в тест-кейсы
    test_cases = []
    for i in range(len(input)):
        test_cases.append({
            "input": input[i].strip(),  # Убираем лишние пробелы
            "expected_output": output[i].strip(),  # Убираем лишние пробелы
        })

    # Создаем новое задание
    new_task = {
        "task_id": generate_task_id(),  # Генерация уникального ID для задания
        "task_name": task_name,
        "task_description": task_description,
        "test_cases": test_cases,
    }

    # Находим блок "Дополнительное"
    additional_block = blocks_collection.find_one({"block_name": "Дополнительное"})

    if additional_block:
        # Если блок "Дополнительное" существует, добавляем задание в него
        blocks_collection.update_one(
            {"block_name": "Дополнительное"},
            {"$push": {"tasks": new_task}}
        )
    else:
        # Если блока "Дополнительное" нет, создаем его
        new_block = {
            "block_id": generate_block_id(),  # Генерация уникального ID для блока
            "block_name": "Дополнительное",
            "tasks": [new_task],  # Добавляем новое задание
        }
        blocks_collection.insert_one(new_block)

    # Перенаправляем пользователя на dashboard
    return RedirectResponse(url="/dashboard", status_code=303)

# Маршрут для отображения страницы добавления ученика
@app.get("/add_student_page", response_class=HTMLResponse)
async def add_student_page(request: Request):
    return templates.TemplateResponse("add_student.html", {"request": request})

# Маршрут для обработки данных формы и добавления ученика
@app.post("/add_student", response_class=RedirectResponse)
async def add_student(
    request: Request,
    group: str = Form(...),
    last_name: str = Form(None),
    first_name: str = Form(None),
    excel_file: UploadFile = File(None),
):
    students = []

    if excel_file:  # Если загружен файл
        try:
            contents = await excel_file.read()
            workbook = load_workbook(io.BytesIO(contents))
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row - индекс строки
                full_name = row[0]  # Индекс столбца
                if full_name:
                    name_parts = full_name.split()
                    last_name = name_parts[0]  # Фамилия
                    first_name = name_parts[1] if len(name_parts) > 1 else ""  # Имя
                    login, password = generate_credentials(last_name, first_name)
                    
                    # Проверка длины логина и пароля
                    if len(login) < MIN_LOGIN_LENGTH or len(login) > MAX_LOGIN_LENGTH:
                        raise HTTPException(status_code=400, detail=f"Длина логина должна быть от {MIN_LOGIN_LENGTH} до {MAX_LOGIN_LENGTH} символов")
                    if len(password) < MIN_PASSWORD_LENGTH or len(password) > MAX_PASSWORD_LENGTH:
                        raise HTTPException(status_code=400, detail=f"Длина пароля должна быть от {MIN_PASSWORD_LENGTH} до {MAX_PASSWORD_LENGTH} символов")
                    
                    students.append({
                        "last_name": last_name,
                        "first_name": first_name,
                        "group": group,
                        "login": login,
                        "password": password,
                    })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Ошибка при обработке файла: {str(e)}")
    else:  # Если данные введены вручную
        if not last_name or not first_name:
            raise HTTPException(status_code=400, detail="Фамилия и имя обязательны")
        login, password = generate_credentials(last_name, first_name)
        
        # Проверка длины логина и пароля
        if len(login) < MIN_LOGIN_LENGTH or len(login) > MAX_LOGIN_LENGTH:
            raise HTTPException(status_code=400, detail=f"Длина логина должна быть от {MIN_LOGIN_LENGTH} до {MAX_LOGIN_LENGTH} символов")
        if len(password) < MIN_PASSWORD_LENGTH or len(password) > MAX_PASSWORD_LENGTH:
            raise HTTPException(status_code=400, detail=f"Длина пароля должна быть от {MIN_PASSWORD_LENGTH} до {MAX_PASSWORD_LENGTH} символов")
        
        students.append({
            "last_name": last_name,
            "first_name": first_name,
            "group": group,
            "login": login,
            "password": password,
        })

    # Добавляем учеников в базу данных
    if students:
        users_collection.insert_many(students)

    # Перенаправляем пользователя на dashboard
    return RedirectResponse(url="/dashboard", status_code=303)

@app.post("/add_lecture", response_class=RedirectResponse)
async def add_lecture_route(
    request: Request,
    title: str = Form(...),
    content_type: List[str] = Form(...),  # Тип контента: text, code, image
    content: List[str] = Form(...),       # Содержимое блока
    code_type: List[str] = Form(default=[]),  # Тип кода (python или пусто для текста/изображения)
):
    # Создаем структурированный контент
    content_blocks = []
    for i in range(len(content_type)):
        block = {
            "type": content_type[i],
            "content": content[i],
        }
        if content_type[i] == "code" and i < len(code_type):
            block["language"] = code_type[i]
        content_blocks.append(block)
    
    # Генерируем уникальный ID для лекции
    lecture_id = generate_lecture_id()
    
    # Создаем документ лекции
    lecture = {
        "lecture_id": lecture_id,
        "title": title,
        "content_blocks": content_blocks,
        "created_at": datetime.now()
    }
    
    # Добавляем лекцию в базу данных
    lectures_collection.insert_one(lecture)
    logger.info(f"Добавлена новая лекция: {title}")
    
    return RedirectResponse(url="/dashboard", status_code=303)

@app.get("/add_lecture_page", response_class=HTMLResponse)
async def add_lecture_page(request: Request):
    return templates.TemplateResponse("add_lecture.html", {"request": request})

@app.get("/lecture/{lecture_id}", response_class=HTMLResponse)
async def show_lecture(request: Request, lecture_id: int):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Получаем лекцию из MongoDB
    lecture = lectures_collection.find_one({"lecture_id": lecture_id})
    if not lecture:
        raise HTTPException(status_code=404, detail="Лекция не найдена")
    
    return templates.TemplateResponse(
        "lecture.html",
        {
            "request": request,
            "lecture": lecture,
            "user": user,
            "is_superuser": user.get("group") == "superuser"
        }
    )
    
@app.get("/lectures", response_class=HTMLResponse)
async def show_lectures(request: Request):
    lectures = get_all_lectures()
    return templates.TemplateResponse(
        "lectures.html",
        {
            "request": request,
            "lectures": lectures
        }
    )
    
# Роут для отображения страницы редактирования лекции
@app.get("/lecture/{lecture_id}/edit", response_class=HTMLResponse)
async def edit_lecture_page(
    request: Request,
    lecture_id: int,
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user or user.get("group") != "superuser":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    
    # Получаем лекцию из MongoDB
    lecture = lectures_collection.find_one({"lecture_id": lecture_id})
    if not lecture:
        raise HTTPException(status_code=404, detail="Лекция не найдена")
    
    return templates.TemplateResponse(
        "edit_lecture.html",  # Шаблон для редактирования лекции
        {
            "request": request,
            "lecture": lecture,
        }
    )

# Роут для обработки данных формы редактирования лекции
@app.post("/lecture/{lecture_id}/edit", response_class=RedirectResponse)
async def edit_lecture(
    request: Request,
    lecture_id: int,
    title: str = Form(...),
    content_type: str = Form(None),  # Делаем поля опциональными
    content: str = Form(None),
    code_type: str = Form(None)
):
    try:
        # Проверяем права доступа
        login = request.session.get("login")
        if not login:
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        user = users_collection.find_one({"login": login})
        if not user or user.get("group") != "superuser":
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        # Проверяем наличие обязательных полей
        if not title:
            raise HTTPException(status_code=422, detail="Заголовок обязателен")
            
        # Преобразуем строки в списки с обработкой ошибок
        content_type_list = []
        content_list = []
        code_type_list = []
        
        if content_type and content:
            try:
                content_type_list = json.loads(content_type)
                content_list = json.loads(content)
                if code_type:
                    code_type_list = json.loads(code_type)
            except json.JSONDecodeError as e:
                logger.error(f"Ошибка при разборе JSON данных: {str(e)}")
                raise HTTPException(status_code=422, detail="Неверный формат данных JSON")
        
        # Проверяем соответствие длин списков
        if len(content_type_list) != len(content_list):
            raise HTTPException(status_code=422, detail="Несоответствие количества типов контента и содержимого")
        
        # Создаем структурированный контент
        content_blocks = []
        for i in range(len(content_type_list)):
            block = {
                "type": content_type_list[i],
                "content": content_list[i],
            }
            if content_type_list[i] == "code":
                block["language"] = code_type_list[i] if i < len(code_type_list) else "python"
            content_blocks.append(block)
        
        # Проверяем существование лекции
        existing_lecture = lectures_collection.find_one({"lecture_id": lecture_id})
        if not existing_lecture:
            logger.error(f"Лекция {lecture_id} не найдена")
            raise HTTPException(status_code=404, detail="Лекция не найдена")
        
        # Обновляем лекцию в MongoDB
        update_result = lectures_collection.update_one(
            {"lecture_id": lecture_id},
            {
                "$set": {
                    "title": title,
                    "content_blocks": content_blocks,
                    "updated_at": datetime.now()
                }
            }
        )
        
        if update_result.modified_count == 0:
            logger.error(f"Не удалось обновить лекцию {lecture_id}")
            raise HTTPException(status_code=500, detail="Не удалось сохранить изменения")
        
        logger.info(f"Лекция {lecture_id} успешно обновлена")
        return RedirectResponse(url=f"/lecture/{lecture_id}", status_code=303)
        
    except HTTPException as he:
        # Перенаправляем обратно на страницу редактирования с сообщением об ошибке
        request.session["error_message"] = str(he.detail)
        return RedirectResponse(url=f"/lecture/{lecture_id}/edit", status_code=303)
    except Exception as e:
        logger.error(f"Ошибка при редактировании лекции {lecture_id}: {str(e)}", exc_info=True)
        request.session["error_message"] = f"Ошибка при сохранении: {str(e)}"
        return RedirectResponse(url=f"/lecture/{lecture_id}/edit", status_code=303)

# Роут для удаления лекции
@app.post("/lecture/{lecture_id}/delete", response_class=RedirectResponse)
async def delete_lecture(
    request: Request,
    lecture_id: int,
):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user or user.get("group") != "superuser":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    
    # Удаляем лекцию из MongoDB
    lectures_collection.delete_one({"lecture_id": lecture_id})
    
    # Перенаправляем пользователя на dashboard
    return RedirectResponse(url="/dashboard", status_code=303)

@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    # Получаем login из сессии
    login = request.session.get("login")
    if not login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем текущего пользователя
    user = users_collection.find_one({"login": login})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    return templates.TemplateResponse(
        "settings.html",
        {
            "request": request,
            "user": user,
            "success_message": request.session.pop("success_message", None),
            "error_message": request.session.pop("error_message", None)
        }
    )

@app.post("/settings/update", response_class=RedirectResponse)
async def update_settings(
    request: Request,
    current_password: str = Form(...),
    new_login: str = Form(None),
    new_password: str = Form(None)
):
    # Получаем текущий login из сессии
    current_login = request.session.get("login")
    if not current_login:
        raise HTTPException(status_code=401, detail="Необходима авторизация")
    
    # Получаем пользователя
    user = users_collection.find_one({"login": current_login})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Проверяем текущий пароль
    if user["password"] != current_password:
        request.session["error_message"] = "Неверный текущий пароль"
        return RedirectResponse(url="/settings", status_code=303)
    
    # Подготавливаем обновления
    updates = {}
    
    # Проверяем новый логин
    if new_login and new_login != current_login:
        # Проверяем длину нового логина
        if len(new_login) < MIN_LOGIN_LENGTH or len(new_login) > MAX_LOGIN_LENGTH:
            request.session["error_message"] = f"Длина логина должна быть от {MIN_LOGIN_LENGTH} до {MAX_LOGIN_LENGTH} символов"
            return RedirectResponse(url="/settings", status_code=303)
            
        # Проверяем, не занят ли новый логин
        if users_collection.find_one({"login": new_login}):
            request.session["error_message"] = "Этот логин уже занят"
            return RedirectResponse(url="/settings", status_code=303)
        updates["login"] = new_login
    
    # Проверяем новый пароль
    if new_password:
        # Проверяем длину нового пароля
        if len(new_password) < MIN_PASSWORD_LENGTH or len(new_password) > MAX_PASSWORD_LENGTH:
            request.session["error_message"] = f"Длина пароля должна быть от {MIN_PASSWORD_LENGTH} до {MAX_PASSWORD_LENGTH} символов"
            return RedirectResponse(url="/settings", status_code=303)
        updates["password"] = new_password
    
    # Если есть что обновлять
    if updates:
        # Обновляем данные пользователя
        users_collection.update_one(
            {"login": current_login},
            {"$set": updates}
        )
        
        # Если логин был изменен, обновляем его в сессии
        if "login" in updates:
            request.session["login"] = new_login
            
            # Обновляем логин в коллекции прогресса
            user_progress_collection.update_many(
                {"login": current_login},
                {"$set": {"login": new_login}}
            )
        
        request.session["success_message"] = "Настройки успешно обновлены"
    
    return RedirectResponse(url="/settings", status_code=303)

@app.post("/run_python")
async def run_python_code(request: Request):
    try:
        # Получаем login из сессии для идентификации пользователя
        login = request.session.get("login")
        if not login:
            return {"error": "Необходима авторизация"}

        # Получаем код из тела запроса
        data = await request.json()
        code = data.get("code")
        input_data = data.get("input", "")
        
        if not code:
            return {"error": "Код не предоставлен"}
        
        # Генерируем уникальный идентификатор для запроса
        request_id = f"{login}_{datetime.now().timestamp()}"
        
        try:
            # Получаем экземпляр CodeTester для текущего пользователя
            tester = await code_tester_manager.get_tester(request_id)
            
            # Выполняем код
            result = await tester.run_code(code, input_data)
            
            # Логируем результат для отладки
            logger.info(f"Результат выполнения кода для запроса {request_id}: {result}")
            
            # Проверяем наличие ошибки
            if result.get("error"):
                return {"error": result["error"]}
            
            # Получаем вывод
            output = result.get("output", "")
            
            # Если есть результат, но нет вывода
            if not output and result.get("result") is not None:
                output = str(result["result"])
                
            return {"output": output or "Код выполнен успешно (нет вывода)"}
        finally:
            # Очищаем ресурсы после выполнения
            await code_tester_manager.cleanup_user(request_id)
            
    except Exception as e:
        logger.error(f"Ошибка при выполнении кода: {str(e)}", exc_info=True)
        return {"error": f"Ошибка выполнения: {str(e)}"}
 
if __name__ == "__main__":
    # Регистрируем обработчики сигналов
    signal.signal(signal.SIGINT, handle_exit)
    signal.signal(signal.SIGTERM, handle_exit)
    
    try:
        import uvicorn
        config = uvicorn.Config(
            app=app,
            host="0.0.0.0",
            port=8000,
            loop="asyncio",
            log_level="info",
            timeout_keep_alive=30,
            timeout_graceful_shutdown=10
        )
        server = uvicorn.Server(config)
        # Запускаем сервер
        asyncio.run(server.serve())
    except KeyboardInterrupt:
        logger.info("Получен сигнал прерывания, начинаем корректное завершение")
    except Exception as e:
        logger.error(f"Неожиданная ошибка: {str(e)}", exc_info=True)
    finally:
        logger.info("Программа завершена")